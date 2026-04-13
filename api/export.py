"""
POST /api/export
Generate a formatted Excel workbook from enriched lead data and return it as a download.
"""
from __future__ import annotations
from datetime import datetime
from io import BytesIO
from typing import Any
import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
app = FastAPI()
COLUMNS: list[tuple[str, int]] = [
    ("No.", 6),
    ("Postcode", 12),
    ("Company Name", 38),
    ("Business Description", 42),
    ("Address", 48),
    ("Phone", 20),
    ("Website", 34),
    ("Estimated Sq Ft", 18),
    ("Size Tier", 13),
]

class LeadRow(BaseModel):
    name: str = ""
    category: str = ""
    address: str = ""
    phone: str = ""
    website: str = ""
    postcode: str = ""
    sqft: int | None = None
    size_tier: str = ""
    description: str = ""

class ExportRequest(BaseModel):
    rows: list[LeadRow]

def create_workbook(rows: list[dict[str, Any]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PE Deal Leads"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    for col_idx, (col_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, row_data in enumerate(rows, start=2):
        values = [
            row_idx - 1,
            row_data.get("postcode", ""),
            row_data.get("name", ""),
            row_data.get("description", ""),
            row_data.get("address", ""),
            row_data.get("phone", ""),
            row_data.get("website", ""),
            row_data.get("sqft"),
            row_data.get("size_tier", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 8 and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    return wb

@app.post("/api/export")
async def export(body: ExportRequest):
    if not body.rows:
        raise HTTPException(status_code=400, detail="No data to export.")
    rows_dicts = [r.model_dump() for r in body.rows]
    wb = create_workbook(rows_dicts)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"PE_Leads_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
