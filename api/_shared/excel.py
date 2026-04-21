"""Excel workbook generation utilities.

Creates consistently-styled Excel workbooks for lead exports, with
professional formatting (frozen header row, alternating row colours, column
auto-filter, and number formatting).
"""
from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Style constants
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid",
)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_ALT_ROW_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid",
)


def create_workbook(
    columns: list[tuple[str, int]],
    rows: list[dict[str, Any]],
    value_extractor: Any,
    *,
    sheet_title: str = "Leads",
    number_format_columns: set[int] | None = None,
) -> openpyxl.Workbook:
    """Create a styled Excel workbook from structured row data.

    Args:
        columns: List of ``(column_name, width)`` tuples defining the header.
        rows: List of dicts containing the row data.
        value_extractor: A callable ``(row_index: int, row_data: dict) -> list``
            that returns the cell values for one row.  *row_index* is 1-based.
        sheet_title: Name for the worksheet tab.
        number_format_columns: Set of 1-based column indices that should be
            formatted as integers with thousands separators (``#,##0``).

    Returns:
        A fully formatted :class:`openpyxl.Workbook`.
    """
    if number_format_columns is None:
        number_format_columns = set()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    _write_header(ws, columns)
    _write_data_rows(ws, columns, rows, value_extractor, number_format_columns)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    return wb


def _write_header(
    ws: Any,
    columns: list[tuple[str, int]],
) -> None:
    """Write and style the header row.

    Args:
        ws: The active worksheet.
        columns: List of ``(column_name, width)`` tuples.
    """
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_data_rows(
    ws: Any,
    columns: list[tuple[str, int]],
    rows: list[dict[str, Any]],
    value_extractor: Any,
    number_format_columns: set[int],
) -> None:
    """Write data rows with alternating shading and number formatting.

    Args:
        ws: The active worksheet.
        columns: Column definitions (used only for length).
        rows: Source data dicts.
        value_extractor: Callable producing cell values per row.
        number_format_columns: 1-based column indices for integer formatting.
    """
    for row_idx, row_data in enumerate(rows, start=2):
        values = value_extractor(row_idx - 1, row_data)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP_ALIGNMENT
            if row_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL
            if col_idx in number_format_columns and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
