#!/usr/bin/env python3
"""
Malaysia Manufacturing Deal Sourcer
====================================
CLI tool for identifying manufacturing/factory targets in Malaysia by postcode.
Uses Google Places API for discovery, OpenStreetMap (Overpass) for building
footprint estimation, and homepage keyword scraping.

Usage:
    python malaysia_sourcer.py --postcodes 40000 40100 40150
    python malaysia_sourcer.py --postcodes 40000 40100 --api-key YOUR_KEY
    python malaysia_sourcer.py --file postcodes.txt
"""
from __future__ import annotations
import argparse
import math
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import googlemaps
import openpyxl
import overpy
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SEARCH_KEYWORDS: list[str] = [
    "factory", "manufacturing", "industrial estate",
    "kilang", "manufacturer", "industrial park",
]
SQ_M_TO_SQ_FT = 10.7639
PLACES_DELAY = 0.25
DETAILS_DELAY = 0.15
OVERPASS_DELAY = 1.0
WEB_SCRAPE_TIMEOUT = 6
WEB_SCRAPE_DELAY = 0.3
HOMEPAGE_KEYWORDS = [
    "employees", "annual capacity", "production capacity",
    "workforce", "headcount", "staff", "revenue",
    "square feet", "square meters", "sq ft", "sq m",
    "ISO 9001", "ISO 14001", "IATF",
]
OVERPASS_RADIUS_M = 80
TIER_SMALL_MAX = 15_000
TIER_MEDIUM_MAX = 50_000


def postcode_to_latlng(gmaps: googlemaps.Client, postcode: str) -> tuple[float, float] | None:
    try:
        results = gmaps.geocode(f"{postcode}, Malaysia")
        if results:
            loc = results[0]["geometry"]["location"]
            return loc["lat"], loc["lng"]
    except Exception as exc:
        print(f"  [WARN] Geocode failed for {postcode}: {exc}")
    return None


def search_places_for_postcode(gmaps, postcode, lat, lng) -> list[dict[str, Any]]:
    seen_ids: set[str] = set()
    results: list[dict[str, Any]] = []
    for keyword in SEARCH_KEYWORDS:
        query = f"{keyword} {postcode} Malaysia"
        try:
            resp = gmaps.places(query=query, location=(lat, lng), radius=5000)
        except Exception as exc:
            print(f"  [WARN] Places search error ({keyword}): {exc}")
            time.sleep(PLACES_DELAY)
            continue
        for place in resp.get("results", []):
            pid = place["place_id"]
            if pid not in seen_ids:
                seen_ids.add(pid)
                results.append(place)
        while resp.get("next_page_token"):
            time.sleep(2)
            try:
                resp = gmaps.places(query=query, page_token=resp["next_page_token"])
                for place in resp.get("results", []):
                    pid = place["place_id"]
                    if pid not in seen_ids:
                        seen_ids.add(pid)
                        results.append(place)
            except Exception:
                break
        time.sleep(PLACES_DELAY)
    return results


def enrich_place(gmaps, place_id: str) -> dict[str, Any]:
    fields = ["name", "formatted_address", "formatted_phone_number", "website", "geometry", "types", "business_status", "url"]
    try:
        detail = gmaps.place(place_id=place_id, fields=fields)
        return detail.get("result", {})
    except Exception as exc:
        print(f"  [WARN] Details fetch failed for {place_id}: {exc}")
        return {}


def _polygon_area_sq_m(coords: list[tuple[float, float]]) -> float:
    n = len(coords)
    if n < 3:
        return 0.0
    ref_lat, ref_lng = coords[0]
    m_per_deg_lat = 111_320.0
    m_per_deg_lng = 111_320.0 * math.cos(math.radians(ref_lat))
    pts = [((lat - ref_lat) * m_per_deg_lat, (lng - ref_lng) * m_per_deg_lng) for lat, lng in coords]
    area = 0.0
    for i in range(n):
        j = (i + 1) % n
        area += pts[i][0] * pts[j][1]
        area -= pts[j][0] * pts[i][1]
    return abs(area) / 2.0


def estimate_building_sqft(lat: float, lng: float) -> float | None:
    api = overpy.Overpass()
    query = f"""
    [out:json][timeout:10];
    (way["building"](around:{OVERPASS_RADIUS_M},{lat},{lng}););
    out body;>;out skel qt;
    """
    try:
        result = api.query(query)
    except Exception as exc:
        print(f"    [WARN] Overpass query failed: {exc}")
        return None
    if not result.ways:
        return None
    best_area = 0.0
    for way in result.ways:
        nodes = way.get_nodes(resolve_missing=False)
        coords = [(float(n.lat), float(n.lon)) for n in nodes if n.lat is not None]
        if len(coords) < 3:
            continue
        area = _polygon_area_sq_m(coords)
        if area > best_area:
            best_area = area
    return best_area * SQ_M_TO_SQ_FT if best_area > 0 else None


def scrape_homepage_keywords(url: str) -> list[str]:
    if not url:
        return []
    try:
        resp = requests.get(url, timeout=WEB_SCRAPE_TIMEOUT, headers={"User-Agent": "Mozilla/5.0 (compatible; DealSourcerBot/1.0)"})
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text(separator=" ", strip=True).lower()
        return [kw for kw in HOMEPAGE_KEYWORDS if kw.lower() in text]
    except Exception:
        return []


def classify_size_tier(sqft: float | None) -> str:
    if sqft is None: return "Unknown"
    if sqft <= TIER_SMALL_MAX: return "Small"
    if sqft <= TIER_MEDIUM_MAX: return "Medium"
    return "Large"


COLUMNS = [
    ("No.", 6), ("Company Name", 38), ("Category / Types", 28),
    ("Full Address", 48), ("Phone", 20), ("Website", 34),
    ("Sq Ft Estimate", 16), ("Size Tier", 13), ("Homepage Keywords", 32), ("Google Maps Link", 42),
]


def write_excel(rows: list[dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manufacturing Leads"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    wrap = Alignment(wrap_text=True, vertical="top")
    for col_idx, (col_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for row_idx, row_data in enumerate(rows, start=2):
        values = [row_idx - 1, row_data.get("name", ""), row_data.get("category", ""), row_data.get("address", ""),
                  row_data.get("phone", ""), row_data.get("website", ""), row_data.get("sqft"),
                  row_data.get("size_tier", ""), row_data.get("homepage_keywords", ""), row_data.get("maps_link", "")]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 7 and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    wb.save(output_path)


def run(postcodes: list[str], api_key: str) -> None:
    gmaps = googlemaps.Client(key=api_key)
    all_leads: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for pc_idx, postcode in enumerate(postcodes, start=1):
        postcode = postcode.strip()
        if not re.fullmatch(r"\d{5}", postcode):
            print(f"[SKIP] Invalid postcode: '{postcode}'")
            continue
        print(f"\n[{pc_idx}/{len(postcodes)}] Processing: {postcode}")
        latlng = postcode_to_latlng(gmaps, postcode)
        if not latlng:
            continue
        lat, lng = latlng
        places = search_places_for_postcode(gmaps, postcode, lat, lng)
        print(f"  Found {len(places)} unique places")
        for i, place in enumerate(places):
            pid = place["place_id"]
            if pid in seen_ids:
                continue
            seen_ids.add(pid)
            print(f"  [{i+1}/{len(places)}] Enriching: {place.get('name', 'Unknown')}")
            detail = enrich_place(gmaps, pid)
            time.sleep(DETAILS_DELAY)
            if not detail:
                continue
            geom = detail.get("geometry", place.get("geometry", {}))
            plat = geom.get("location", {}).get("lat", lat)
            plng = geom.get("location", {}).get("lng", lng)
            sqft = estimate_building_sqft(plat, plng)
            time.sleep(OVERPASS_DELAY)
            website = detail.get("website", "")
            kw_found = scrape_homepage_keywords(website) if website else []
            if website:
                time.sleep(WEB_SCRAPE_DELAY)
            types_list = detail.get("types", place.get("types", []))
            all_leads.append({
                "name": detail.get("name", place.get("name", "")),
                "category": ", ".join(t.replace("_", " ").title() for t in types_list if t != "establishment"),
                "address": detail.get("formatted_address", ""),
                "phone": detail.get("formatted_phone_number", ""),
                "website": website,
                "sqft": round(sqft) if sqft else None,
                "size_tier": classify_size_tier(sqft),
                "homepage_keywords": ", ".join(kw_found),
                "maps_link": detail.get("url", ""),
            })
    deduped: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    for lead in all_leads:
        key = (lead["name"].lower().strip(), lead["address"].lower().strip())
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(lead)
    deduped.sort(key=lambda r: (r["sqft"] is None, -(r["sqft"] or 0), r["name"]))
    if not deduped:
        print("\nNo leads found.")
        return
    output_path = Path.cwd() / f"Source_Leads_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    write_excel(deduped, output_path)
    print(f"\nComplete: {len(deduped)} leads → {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Malaysia Manufacturing Deal Sourcer")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--postcodes", "-p", nargs="+")
    group.add_argument("--file", "-f", type=str)
    parser.add_argument("--api-key", "-k", type=str, default=None)
    args = parser.parse_args()
    api_key = args.api_key or os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if not api_key:
        print("ERROR: Provide a Google Maps API key via --api-key or GOOGLE_MAPS_API_KEY env var.")
        sys.exit(1)
    if args.file:
        path = Path(args.file)
        if not path.exists():
            print(f"ERROR: File not found: {path}")
            sys.exit(1)
        postcodes = [l.strip() for l in path.read_text().splitlines() if l.strip() and not l.strip().startswith("#")]
    else:
        postcodes = args.postcodes
    run(postcodes, api_key)


if __name__ == "__main__":
    main()
